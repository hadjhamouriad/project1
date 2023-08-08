import kivy
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from kivy.uix.popup import Popup
from kivy.uix.scrollview import ScrollView
from kivy.uix.gridlayout import GridLayout
from kivy.uix.image import Image  # Moved this import up
import requests
from openpyxl import Workbook, load_workbook
from datetime import datetime
import schedule
import threading
import time
import matplotlib.pyplot as plt
import numpy as np

kivy.require('1.11.1')
class GraphPopup(Popup):
    def __init__(self, cities, temps, graph_image_path, **kwargs):
        super().__init__(**kwargs)
        layout = BoxLayout(orientation='vertical')

        # Création du graphique
        fig, ax = plt.subplots(figsize=(8, 6))
        for i in range(len(cities)):
            ax.plot(temps[i], label=cities[i])

        ax.set_xlabel("Index de données")
        ax.set_ylabel("Température (°C)")
        ax.set_title("Évolution de la température par ville")
        ax.legend()
        ax.grid(True)

        graph_image_path = graph_image_path + ".png"
        fig.savefig(graph_image_path)

        graph_image = Image(source=graph_image_path, size_hint=(1, None), height=500)
        layout.add_widget(graph_image)

        close_button = Button(text="Fermer", size_hint_y=None, height=50)
        close_button.bind(on_press=self.dismiss)
        layout.add_widget(close_button)

        self.content = layout

class WeatherApp(App):
    def build(self):
        layout = BoxLayout(orientation='vertical')

        self.city_input = TextInput(hint_text="Entrez le nom de la ville")
        self.weather_label = Label(text="")
        self.start_search_button = Button(text="Démarrer les recherches automatiques", on_press=self.start_auto_updates)
        self.stop_search_button = Button(text="Arrêter les recherches automatiques", on_press=self.stop_auto_search, disabled=True)
        self.history_button = Button(text="Historique", on_press=self.show_history_popup)
        self.statistics_button = Button(text="Statistiques", on_press=self.show_temperature_graph)
        layout.add_widget(self.city_input)
        layout.add_widget(self.weather_label)
        layout.add_widget(self.start_search_button)
        layout.add_widget(self.stop_search_button)
        layout.add_widget(self.history_button)
        layout.add_widget(self.statistics_button)

        self.popup = None
        return layout

    def get_weather(self, city):
        api_key = "b5ecc8297ad7e069460276ac17dab215"

        url = f"http://api.openweathermap.org/data/2.5/weather?q={city}&appid={api_key}&units=metric"
        response = requests.get(url)

        if response.status_code == 200:
            weather_data = response.json()
            main_weather = weather_data["weather"][0]["main"]
            temp = weather_data["main"]["temp"]

            weather_info = f"Conditions météorologiques à {city} : {main_weather}\nTempérature : {temp} °C"
            self.weather_label.text = weather_info

            self.save_weather_to_excel(city, main_weather, temp)
        else:
            self.weather_label.text = "Erreur lors de la récupération des données météorologiques"

    def save_weather_to_excel(self, city, main_weather, temp):
        try:
            wb = load_workbook(r"C:\Users\riad\Desktop\weather_history.xlsx")
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.append(["Date", "Heure", "Ville", "Conditions météo", "Température"])

        ws.append([datetime.now().strftime("%Y-%m-%d"), datetime.now().strftime("%H:%M:%S"), city, main_weather, temp])

        wb.save(r"C:\Users\riad\Desktop\weather_history.xlsx")

    def scheduled_search(self, city):
        self.get_weather(city)

    def start_auto_updates(self, instance):
        city = self.city_input.text
        if city:
            self.get_weather(city)
            self.stop_search_button.disabled = False
            self.start_search_button.disabled = True
            schedule.every(5).minutes.do(self.scheduled_search, city)
            threading.Thread(target=self.run_schedule).start()

    def stop_auto_search(self, instance):
        schedule.clear()
        self.stop_search_button.disabled = True
        self.start_search_button.disabled = False

    def show_history_popup(self, instance):
        history_popup_content = BoxLayout(orientation='vertical')

        # Grille pour afficher les données de l'historique
        history_grid = GridLayout(cols=5, spacing=20, padding=15, size_hint_y=None)
        history_grid.bind(minimum_height=history_grid.setter('height'))

        try:
            wb = load_workbook(r"C:\Users\riad\Desktop\weather_history.xlsx")
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                date = row[0]
                time = row[1]
                city = row[2]
                main_weather = row[3]
                temp = row[4]

                history_grid.add_widget(Label(text=str(date)))
                history_grid.add_widget(Label(text=str(time)))
                history_grid.add_widget(Label(text=str(city)))
                history_grid.add_widget(Label(text=str(main_weather)))
                history_grid.add_widget(Label(text=str(temp)))

        except FileNotFoundError:
            history_grid.add_widget(Label(text="Aucune donnée disponible.", col_span=5))

        scroll_view = ScrollView()
        scroll_view.add_widget(history_grid)
        history_popup_content.add_widget(scroll_view)

        # Ajouter le bouton de suppression de l'historique
        clear_button = Button(text="Supprimer l'historique", size_hint=(None, None), size=(200, 50),
                              on_press=self.clear_history)
        history_popup_content.add_widget(clear_button)

        # Ajouter le bouton de retour au menu principal
        back_button = Button(text="Retour au menu principal", size_hint=(None, None), size=(200, 50),
                             on_press=self.close_history_popup)
        history_popup_content.add_widget(back_button)

        # Créer et afficher la fenêtre contextuelle de l'historique
        self.popup = Popup(title='Historique des Données', content=history_popup_content, size_hint=(None, None),
                           size=(800, 600))
        self.popup.open()

    def clear_history(self, instance):
        try:
            wb = load_workbook(r"C:\Users\riad\Desktop\weather_history.xlsx")
            ws = wb.active
            ws.delete_rows(2, ws.max_row - 1)
            wb.save(r"C:\Users\riad\Desktop\weather_history.xlsx")
            # Only dismiss the popup if the file is successfully cleared
            self.popup.dismiss()
        except FileNotFoundError:
            pass
    def close_history_popup(self, instance):
        self.popup.dismiss()

    def run_schedule(self):
        while True:
            schedule.run_pending()
            time.sleep(1)

    def show_temperature_graph(self, instance):
        try:
            wb = load_workbook(r"C:\Users\riad\Desktop\weather_history.xlsx")
            ws = wb.active

            cities = []
            temps = []

            for row in ws.iter_rows(min_row=2, values_only=True):
                city = row[2]
                temp = row[4]

                if city not in cities:
                    cities.append(city)
                    temps.append([temp])
                else:
                    index = cities.index(city)
                    temps[index].append(temp)

            graph_image_path = "temperature_graph"
            # Création du graphique
            plt.figure(figsize=(10, 6))
            for i in range(len(cities)):
                plt.plot(temps[i], label=cities[i])

            plt.xlabel("Index de données")
            plt.ylabel("Température (°C)")
            plt.title("Évolution de la température par ville")
            plt.legend()
            plt.grid(True)

            graph_popup = GraphPopup(
                title="Statistiques de température par ville",
                cities=cities,
                temps=temps,
                graph_image_path=graph_image_path,
            )
            plt.savefig(graph_image_path + ".png")
            graph_popup.open()

        except FileNotFoundError:
            print("Aucune donnée disponible.")




if __name__ == '__main__':
    WeatherApp().run()
